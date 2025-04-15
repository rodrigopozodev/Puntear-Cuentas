# src/api.py

from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import shutil, os, subprocess
from pathlib import Path
from typing import Optional
import openpyxl
import uvicorn

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
INFORMES_DIR = BASE_DIR / "informes"
SRC_DIR = BASE_DIR / "src"

DATA_DIR.mkdir(parents=True, exist_ok=True)
INFORMES_DIR.mkdir(parents=True, exist_ok=True)

@app.post("/save-excel")
async def save_excel(file: UploadFile = File(...)):
    try:
        file_path = DATA_DIR / file.filename
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"message": "File uploaded successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/execute-python")
async def execute_python():
    try:
        process = subprocess.run(
            ["python", str(SRC_DIR / "main.py")],
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True
        )
        if process.returncode == 0:
            return {"success": True, "output": process.stdout}
        else:
            raise HTTPException(status_code=500, detail=process.stderr)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/informes")
async def list_informes():
    files_data = []
    for root, _, files in os.walk(INFORMES_DIR):
        for f in files:
            if f.endswith((".xlsx", ".xls")):
                path = Path(root) / f
                files_data.append({
                    "name": f,
                    "path": str(path),
                    "folder": str(Path(root).relative_to(INFORMES_DIR)),
                    "size": path.stat().st_size,
                    "createdAt": path.stat().st_ctime
                })
    return files_data

@app.get("/informes/download")
async def download_file(path: str = Query(...)):
    file_path = Path(path)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(path=file_path, filename=file_path.name, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get("/informes/content")
async def read_excel(path: str = Query(...)):
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows = [
            {headers[i]: cell.value for i, cell in enumerate(row)}
            for row in sheet.iter_rows(min_row=2)
        ]
        return {"headers": headers, "rows": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al leer archivo: {e}")

@app.delete("/informes")
async def delete_all_informes():
    try:
        for file in INFORMES_DIR.iterdir():
            if file.is_file():
                file.unlink()
        return {"message": "Informes borrados exitosamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/delete-informes-folder")
async def reset_informes_folder():
    try:
        if INFORMES_DIR.exists():
            shutil.rmtree(INFORMES_DIR)
        INFORMES_DIR.mkdir(parents=True, exist_ok=True)
        return {"message": "Carpeta de informes reiniciada correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Para desarrollo local
if __name__ == "__main__":
    uvicorn.run("src.api:app", host="127.0.0.1", port=3000, reload=True)
