from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from tqdm import tqdm

def generar_archivo_excel(df_datos, archivo_plantilla, archivo_salida):
    """Genera un archivo Excel manteniendo el formato de la plantilla original."""
    # Cargamos la plantilla original para mantener su formato
    wb = load_workbook(archivo_plantilla)
    ws = wb.active
    
    print(f"\nGenerando archivo: {archivo_salida}")
    total_filas = len(df_datos) + 1  # +1 por el encabezado
    
    with tqdm(total=total_filas, desc="Escribiendo datos") as pbar:
        # Escribimos los encabezados
        for col_idx, column in enumerate(df_datos.columns, 1):
            ws.cell(row=1, column=col_idx, value=column)
        pbar.update(1)
        
        # Escribimos los datos manteniendo el formato
        for idx, row in df_datos.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=idx + 2, column=col_idx, value=value)
            pbar.update(1)
    
    # Eliminamos las filas sobrantes si existen
    if ws.max_row > len(df_datos) + 1:
        ws.delete_rows(len(df_datos) + 2, ws.max_row - (len(df_datos) + 1))
    
    wb.save(archivo_salida)