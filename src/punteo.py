import os
import pandas as pd
import numpy as np
from itertools import combinations
from tqdm import tqdm
from utils import cargar_datos, crear_directorio
from openpyxl import load_workbook

def emparejar_iguales(df):
    """Empareja valores exactamente iguales con barra de progreso."""
    df['Indice_Punteo'] = None
    punteo_index = 1
    
    print("Construyendo índices de búsqueda...")
    debe_dict = {}
    haber_dict = {}
    
    for idx, row in df.iterrows():
        debe = round(row['Debe'], 2)
        haber = round(row['Haber'], 2)
        
        if debe > 0:
            if debe not in debe_dict:
                debe_dict[debe] = []
            debe_dict[debe].append(idx)
                
        if haber > 0:
            if haber not in haber_dict:
                haber_dict[haber] = []
            haber_dict[haber].append(idx)
    
    print("Buscando coincidencias exactas...")
    valores_debe = set(debe_dict.keys())
    valores_haber = set(haber_dict.keys())
    valores_comunes = valores_debe.intersection(valores_haber)
    
    for valor in tqdm(valores_comunes, desc="Procesando valores"):
        debe_indices = debe_dict[valor]
        haber_indices = haber_dict[valor]
        
        for debe_idx in debe_indices:
            if df.at[debe_idx, 'Indice_Punteo'] is not None:
                continue
            
            for haber_idx in haber_indices:
                if (df.at[haber_idx, 'Indice_Punteo'] is None and 
                    debe_idx != haber_idx):
                    df.loc[[debe_idx, haber_idx], 'Indice_Punteo'] = punteo_index
                    punteo_index += 1
                    break
    
    return df, punteo_index

def emparejar_por_suma(df, punteo_index, max_combinaciones=10, tolerancia=0):
    """Busca combinaciones exactas de sumas con optimización de rendimiento."""
    no_punteados = df[df['Indice_Punteo'].isna()].copy()
    
    print(f"Buscando combinaciones de hasta {max_combinaciones} registros...")
    debe_valores = no_punteados[no_punteados['Debe'] > 0].sort_values('Debe', ascending=False)
    haber_valores = no_punteados[no_punteados['Haber'] > 0].sort_values('Haber', ascending=False)
    
    debe_array = debe_valores['Debe'].values
    debe_indices = debe_valores.index.values
    
    # Creamos un diccionario para almacenar sumas precalculadas
    suma_cache = {}
    
    def calcular_sumas_para_n(valores, n):
        """Precalcula sumas para n elementos y las almacena en caché."""
        key = f"{len(valores)}_{n}"
        if key not in suma_cache:
            sumas = {}
            for comb_indices in combinations(range(len(valores)), n):
                suma = round(sum(valores[list(comb_indices)]), 2)
                if suma not in sumas:
                    sumas[suma] = []
                sumas[suma].append(comb_indices)
            suma_cache[key] = sumas
        return suma_cache[key]
    
    # Procesamos por bloques para optimizar memoria
    chunk_size = 1000
    for chunk_start in range(0, len(haber_valores), chunk_size):
        chunk_end = min(chunk_start + chunk_size, len(haber_valores))
        chunk_haber = haber_valores.iloc[chunk_start:chunk_end]
        
        for _, haber_fila in tqdm(chunk_haber.iterrows(),
                                 desc=f"Procesando bloque {chunk_start//chunk_size + 1}",
                                 total=len(chunk_haber)):
            objetivo = round(haber_fila['Haber'], 2)
            if objetivo == 0:
                continue
            
            # Filtramos valores mayores al objetivo
            mask = debe_array <= objetivo * 1.0001  # Pequeño margen para redondeo
            valores_filtrados = debe_array[mask]
            indices_filtrados = debe_indices[mask]
            
            if len(valores_filtrados) < 2:
                continue
            
            # Optimizamos la búsqueda limitando candidatos
            max_candidates = min(20 if max_combinaciones <= 5 else 15, len(valores_filtrados))
            valores_filtrados = valores_filtrados[:max_candidates]
            indices_filtrados = indices_filtrados[:max_candidates]
            
            # Búsqueda optimizada por número de combinaciones
            for n in range(2, min(max_combinaciones + 1, len(valores_filtrados) + 1)):
                # Para combinaciones grandes, solo buscamos si hay suficiente probabilidad
                if n > 5 and len(valores_filtrados) < n * 2:
                    continue
                
                # Optimización: primero verificamos si es posible alcanzar el objetivo
                min_suma = sum(valores_filtrados[-n:])
                max_suma = sum(valores_filtrados[:n])
                if not (min_suma <= objetivo <= max_suma):
                    continue
                
                # Usamos la caché de sumas precalculadas
                sumas = calcular_sumas_para_n(valores_filtrados, n)
                if objetivo in sumas:
                    for comb_indices in sumas[objetivo]:
                        indices = indices_filtrados[list(comb_indices)]
                        if df.loc[indices, 'Indice_Punteo'].isna().all():
                            df.loc[indices, 'Indice_Punteo'] = punteo_index
                            df.at[haber_fila.name, 'Indice_Punteo'] = punteo_index
                            punteo_index += 1
                            
                            # Actualizamos arrays de búsqueda
                            mask = ~np.isin(range(len(valores_filtrados)), comb_indices)
                            valores_filtrados = valores_filtrados[mask]
                            indices_filtrados = indices_filtrados[mask]
                            break
    
    # Limpiamos la caché
    suma_cache.clear()
    return df

def generar_informes(df, archivo):
    """Genera los informes con los datos punteados y no punteados, conservando el formato original."""
    # Filtramos los datos
    emparejados = df[df['Indice_Punteo'].notna()]  # Solo filas con Indice_Punteo no nulo
    no_emparejados = df[df['Indice_Punteo'].isna()]  # Solo filas con Indice_Punteo nulo

    # Configuramos las rutas
    nombre_base = os.path.basename(archivo).replace(".xlsx", "")
    carpeta_informes = "informes"
    subcarpeta_informes = os.path.join(carpeta_informes, nombre_base)
    
    # Creamos las carpetas necesarias
    crear_directorio(carpeta_informes)
    crear_directorio(subcarpeta_informes)

    # Cargamos el archivo original con openpyxl
    wb = load_workbook(archivo)
    ws = wb.active

    # Verificamos si la columna 'Indice_Punteo' ya existe
    encabezados = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    if 'Indice_Punteo' not in encabezados:
        # Añadimos el título de la columna 'Indice_Punteo' en la última columna
        ws.cell(row=1, column=ws.max_column + 1, value='Indice_Punteo')

    # Limpiamos la hoja activa antes de escribir los datos punteados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Sobrescribimos los datos procesados en la hoja original (solo emparejados)
    fila_actual = 2  # Comenzamos en la fila 2 (después de los encabezados)
    for _, row in emparejados.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=fila_actual, column=col_idx, value=value)
        fila_actual += 1  # Avanzamos a la siguiente fila

    # Guardamos el archivo con los datos punteados
    archivo_punteado = f"{subcarpeta_informes}/{nombre_base}_punteado.xlsx"
    wb.save(archivo_punteado)

    # Crear un nuevo archivo para los no emparejados
    archivo_no_punteado = f"{subcarpeta_informes}/{nombre_base}_no_punteado.xlsx"
    wb_no_punteado = load_workbook(archivo)  # Cargamos el archivo original para conservar el formato
    ws_no_punteado = wb_no_punteado.active

    # Limpiamos las filas existentes en la hoja activa
    for row in ws_no_punteado.iter_rows(min_row=2, max_row=ws_no_punteado.max_row):
        for cell in row:
            cell.value = None

    # Escribimos los datos no emparejados en el nuevo archivo
    fila_actual = 2  # Comenzamos en la fila 2 (después de los encabezados)
    for _, row in no_emparejados.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws_no_punteado.cell(row=fila_actual, column=col_idx, value=value)
        fila_actual += 1  # Avanzamos a la siguiente fila

    # Guardamos el archivo de no emparejados
    wb_no_punteado.save(archivo_no_punteado)

    print(f"Informes generados para {archivo}:")
    print(f"→ Archivo punteado: {archivo_punteado}")
    print(f"→ Archivo no punteado: {archivo_no_punteado}")
